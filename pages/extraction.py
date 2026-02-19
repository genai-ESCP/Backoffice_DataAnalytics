import io
import re
import unicodedata

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils.auth import require_auth
from utils.ui import apply_dashboard_style, card, render_sidebar

COURSE_2425 = "2425ALL_OL_GENAI_00"
COURSE_2526_FALL_NEW = "2526ALL_OL_GENAI_00"
COURSE_2526_FALL_RETAKE = "2526ALL_OL_GENAI_02"
COURSE_2526_SPRING_NEW = "2526ALL_SPR_GENAI_00"
COURSE_2526_SPRING_RETAKE = "2526ALL_SPR_GENAI_02"

COURSE_PRESETS = [
    COURSE_2425,
    COURSE_2526_FALL_NEW,
    COURSE_2526_FALL_RETAKE,
    COURSE_2526_SPRING_NEW,
    COURSE_2526_SPRING_RETAKE,
]

DEFAULT_COURSE_PRESET = COURSE_2526_SPRING_NEW
DEFAULT_SEJI_SHEET = f"gc_{DEFAULT_COURSE_PRESET}_fullgc_2"
DEFAULT_OUTPUT_XLSX = "Data_GenAI_2526ALL_SPR_GENAI_00.xlsx"

SEJI_LASTNAME_COL = "Last Name"
SEJI_FIRSTNAME_COL = "First Name"
SEJI_USERNAME_COL = "Username"
SEJI_DELETE_COLS = ["Student ID", "Last Access", "Availability"]

SEJI_STUDENT_ID_COL = "Student ID"
SEJI_HOURS_OUT = "Hours in Course"
SEJI_EMAIL_OUT = "Email"
SEJI_VERDICT_OUT = "Verdict"

H_CODE_COL = "Code d'étudiant"
H_FIRST_COL = "Prénom"
H_LAST_COL = "Nom"
H_EMAIL_COL = "Adresse e-mail"
H_HOURS_COL = "Temps passé dans le cours (en heures)"

FINAL_COL_ORDER = [
    "Last Name",
    "First Name",
    "Student ID",
    "Email",
    "Verdict",
    "Hours in Course",
    "Overall Grade",
]

_HEADER_STRIP_RE = re.compile(r"\s*\[.*?\]\s*(?:\|\s*\d+)?\s*$", flags=re.UNICODE)


def norm_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


def strip_accents(s):
    s = unicodedata.normalize("NFKD", norm_str(s))
    return "".join(ch for ch in s if not unicodedata.combining(ch))


def norm_key(x):
    return strip_accents(norm_str(x)).lower()


def norm_id(x):
    return norm_str(x)


def first_nonempty(series):
    for v in series:
        if isinstance(v, str) and v.strip():
            return v.strip().lower()
    return ""


def clean_header_text(h):
    if h is None:
        return ""
    h = norm_str(h)
    h = _HEADER_STRIP_RE.sub("", h)
    h = re.sub(r"\s+", " ", h).strip()
    return h


def clean_headers_in_df(df):
    cols = list(df.columns)
    cleaned = [clean_header_text(c) for c in cols]
    seen = {}
    final = []
    for original, c in zip(cols, cleaned):
        base = c or original or "Unnamed"
        key = base
        if key in seen:
            seen[key] += 1
            key = f"{base}_{seen[key]}"
        else:
            seen[key] = 0
        final.append(key)
    df2 = df.copy()
    df2.columns = final
    return df2


def read_gradebook_maybe_fake_xls(uploaded_file):
    name = uploaded_file.name.lower()
    raw = uploaded_file.getvalue()
    bio = io.BytesIO(raw)

    if name.endswith(".xlsx"):
        xl = pd.ExcelFile(bio, engine="openpyxl")
        return {sh: pd.read_excel(io.BytesIO(raw), sheet_name=sh, engine="openpyxl") for sh in xl.sheet_names}

    if name.endswith(".xls"):
        try:
            xl = pd.ExcelFile(bio, engine="xlrd")
            return {sh: pd.read_excel(io.BytesIO(raw), sheet_name=sh, engine="xlrd") for sh in xl.sheet_names}
        except Exception:
            df = pd.read_csv(io.BytesIO(raw), sep="\t", encoding="utf-16", engine="python")
            return {"Sheet1": df}

    xl = pd.ExcelFile(bio)
    return {sh: pd.read_excel(io.BytesIO(raw), sheet_name=sh) for sh in xl.sheet_names}


def read_excel_flexible(uploaded_file, sheet_name=None, **kwargs):
    name = uploaded_file.name.lower()
    raw = uploaded_file.getvalue()
    if name.endswith(".xls"):
        try:
            return pd.read_excel(io.BytesIO(raw), sheet_name=sheet_name, engine="xlrd", **kwargs)
        except Exception as e:
            raise RuntimeError(
                f"Failed to read .xls file '{uploaded_file.name}'. Install 'xlrd' for true .xls support. Error: {e}"
            )
    return pd.read_excel(io.BytesIO(raw), sheet_name=sheet_name, **kwargs)


def process_uploaded_files(seji_file, hours_file, output_name, seji_sheet):
    seji_sheets = read_gradebook_maybe_fake_xls(seji_file)

    if seji_sheet not in seji_sheets:
        if "Sheet1" in seji_sheets:
            seji_sheet = "Sheet1"
        else:
            raise ValueError(f"Sheet '{seji_sheet}' not found. Available: {list(seji_sheets.keys())}")

    seji = seji_sheets[seji_sheet].copy()
    seji = clean_headers_in_df(seji)

    for c in SEJI_DELETE_COLS:
        if c in seji.columns:
            seji.drop(columns=[c], inplace=True)

    user_col_candidates = [c for c in seji.columns if norm_key(c) == norm_key(SEJI_USERNAME_COL)]
    if not user_col_candidates:
        user_col_candidates = [c for c in seji.columns if "username" in norm_key(c)]
    if not user_col_candidates:
        raise ValueError(
            f"Column '{SEJI_USERNAME_COL}' not found in SEJI sheet. Available columns: {list(seji.columns)}"
        )
    seji.rename(columns={user_col_candidates[0]: SEJI_STUDENT_ID_COL}, inplace=True)

    hours_raw = read_excel_flexible(hours_file, sheet_name=0, skiprows=3)
    if isinstance(hours_raw, dict):
        first_sheet = next(iter(hours_raw))
        hours_raw = hours_raw[first_sheet]
    hours_raw = clean_headers_in_df(hours_raw)

    def _tokens(text):
        return [t for t in re.split(r"[^a-z0-9]+", norm_key(text)) if t]

    def find_col_by_keyword(df_cols, keywords, exclude=None):
        exclude = exclude or set()
        keys = [norm_key(k) for k in df_cols]
        kws = [norm_key(kw) for kw in keywords]
        for idx, k in enumerate(keys):
            col = df_cols[idx]
            if col in exclude:
                continue
            tok = set(_tokens(k))
            if all(kw in tok for kw in kws):
                return col
        return None

    def find_col_by_any_keyword_group(df_cols, keyword_groups, exclude=None):
        for group in keyword_groups:
            found = find_col_by_keyword(df_cols, group, exclude=exclude)
            if found:
                return found
        return None

    h_code = find_col_by_any_keyword_group(hours_raw.columns, [["code", "etudiant"], ["code", "student"]]) or H_CODE_COL
    h_first = find_col_by_any_keyword_group(hours_raw.columns, [["prenom"], ["first"]], exclude={h_code}) or H_FIRST_COL
    h_last = find_col_by_any_keyword_group(hours_raw.columns, [["nom"], ["last"]], exclude={h_code, h_first}) or H_LAST_COL
    h_email = (
        find_col_by_any_keyword_group(hours_raw.columns, [["adresse", "email"], ["email"]], exclude={h_code, h_first, h_last})
        or H_EMAIL_COL
    )
    h_hours = (
        find_col_by_any_keyword_group(hours_raw.columns, [["temps", "heure"], ["hours"], ["time", "spent"]], exclude={h_code, h_first, h_last, h_email})
        or H_HOURS_COL
    )

    needed_cols = [h_code, h_first, h_last, h_email, h_hours]
    if len(set(needed_cols)) != len(needed_cols):
        raise ValueError(
            "HOURS auto-detection picked overlapping columns. "
            f"Detected: code={h_code}, first={h_first}, last={h_last}, email={h_email}, hours={h_hours}. "
            "Please verify headers in the uploaded HOURS file."
        )
    missing = [c for c in needed_cols if c not in hours_raw.columns]
    if missing:
        raise ValueError(f"Missing columns in HOURS file (after auto-detection): {missing}\nFound: {list(hours_raw.columns)}")

    hours = hours_raw[[h_code, h_first, h_last, h_email, h_hours]].copy()
    hours["_id"] = hours[h_code].map(norm_id)
    hours["_name"] = (hours[h_first].map(norm_key) + "|" + hours[h_last].map(norm_key)).map(norm_str)
    hours["_hours"] = pd.to_numeric(hours[h_hours], errors="coerce")
    hours["_email"] = hours[h_email].map(lambda x: norm_str(x).lower())

    hours_by_id = (
        hours.groupby("_id", as_index=False).agg(
            _hours=("_hours", "max"),
            _email=("_email", first_nonempty),
        )
    )

    hours_by_name = (
        hours.groupby("_name", as_index=False).agg(
            _hours_n=("_hours", "max"),
            _email_n=("_email", first_nonempty),
        )
    )

    seji["_id"] = seji[SEJI_STUDENT_ID_COL].map(norm_id)
    seji["_name"] = (seji[SEJI_FIRSTNAME_COL].map(norm_key) + "|" + seji[SEJI_LASTNAME_COL].map(norm_key)).map(norm_str)

    merged = seji.merge(hours_by_id, on="_id", how="left")
    merged = merged.merge(hours_by_name, on="_name", how="left")

    merged[SEJI_HOURS_OUT] = merged.get("_hours")
    merged.loc[merged[SEJI_HOURS_OUT].isna(), SEJI_HOURS_OUT] = merged.loc[merged[SEJI_HOURS_OUT].isna(), "_hours_n"]
    merged[SEJI_HOURS_OUT] = merged[SEJI_HOURS_OUT].fillna(0)

    merged[SEJI_EMAIL_OUT] = merged.get("_email")
    merged.loc[
        merged[SEJI_EMAIL_OUT].isna() | (merged[SEJI_EMAIL_OUT] == ""),
        SEJI_EMAIL_OUT,
    ] = merged.loc[
        merged[SEJI_EMAIL_OUT].isna() | (merged[SEJI_EMAIL_OUT] == ""),
        "_email_n",
    ]
    merged[SEJI_EMAIL_OUT] = merged[SEJI_EMAIL_OUT].fillna("")

    if SEJI_VERDICT_OUT not in merged.columns:
        merged[SEJI_VERDICT_OUT] = ""

    merged.drop(columns=[c for c in ["_id", "_name", "_hours", "_email", "_hours_n", "_email_n"] if c in merged.columns], inplace=True)
    merged = clean_headers_in_df(merged)

    desired_clean = [clean_header_text(c) for c in FINAL_COL_ORDER]
    final_cols = [c for c in desired_clean if c in merged.columns]
    extras = [c for c in merged.columns if c not in final_cols]
    merged = merged[final_cols + extras]

    out_buffer = io.BytesIO()
    with pd.ExcelWriter(out_buffer, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name=seji_sheet, index=False)
        for sh, df in seji_sheets.items():
            if sh == seji_sheet:
                continue
            clean_headers_in_df(df).to_excel(writer, sheet_name=sh, index=False)
    out_buffer.seek(0)

    wb = load_workbook(out_buffer)
    ws = wb[seji_sheet]
    headers = {cell.value: cell.column for cell in ws[1] if cell.value is not None}

    verdict_col = headers.get(SEJI_VERDICT_OUT)
    hours_col = headers.get(SEJI_HOURS_OUT)

    overall_col = None
    for hname, col in headers.items():
        lk = norm_key(hname)
        if "overall" in lk and "grade" in lk:
            overall_col = col
            break
    if overall_col is None:
        for hname, col in headers.items():
            lk = norm_key(hname)
            if "overall" in lk or "grade" in lk:
                overall_col = col
                break

    if not (verdict_col and hours_col and overall_col):
        raise ValueError(
            "Could not locate one of the required headers for formula injection: "
            f"Verdict={verdict_col}, Hours in Course={hours_col}, Overall Grade={overall_col}. "
            "Headers found: " + ", ".join([str(k) for k in headers.keys()])
        )

    v_col = get_column_letter(verdict_col)
    f_col = get_column_letter(hours_col)
    g_col = get_column_letter(overall_col)

    for r in range(2, ws.max_row + 1):
        ws[f"{v_col}{r}"].value = f'=IF(AND({f_col}{r}>=0.5, {g_col}{r}>=50), "Passed", "Failed")'

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output.getvalue(), output_name, ws.max_row - 1


st.set_page_config(page_title="Extraction", page_icon="E", layout="wide", initial_sidebar_state="expanded")
apply_dashboard_style()
require_auth()
render_sidebar("Extraction")

st.markdown("# Extraction")
st.markdown(
    "<div class='small-muted'>Merge a Blackboard gradebook with a Hours export and generate a ready-to-download Excel file.</div>",
    unsafe_allow_html=True,
)

col1, col2 = st.columns([1, 1], gap="large")
with col1:
    seji_file = st.file_uploader("Gradebook file (`gc_*.xls` or `gc_*.xlsx`)", type=["xls", "xlsx"])
with col2:
    hours_file = st.file_uploader("Hours file (`.xlsx`)", type=["xlsx"])

cfg1, cfg2 = st.columns([1, 1], gap="large")
with cfg1:
    selected_course_preset = st.selectbox(
        "Course preset",
        options=COURSE_PRESETS,
        index=COURSE_PRESETS.index(DEFAULT_COURSE_PRESET),
    )
    seji_sheet = st.text_input("Target sheet name", value=f"gc_{selected_course_preset}_fullgc_2")
with cfg2:
    output_name = st.text_input("Output filename", value=f"Data_GenAI_{selected_course_preset}.xlsx")

if st.button("Process extraction", type="primary", use_container_width=True):
    if seji_file is None or hours_file is None:
        st.error("Upload both files before processing.")
    else:
        with st.spinner("Processing files..."):
            try:
                output_bytes, output_filename, row_count = process_uploaded_files(
                    seji_file=seji_file,
                    hours_file=hours_file,
                    output_name=output_name.strip() or DEFAULT_OUTPUT_XLSX,
                    seji_sheet=seji_sheet.strip() or DEFAULT_SEJI_SHEET,
                )
                st.success(f"Done. Generated file with {row_count} rows.")
                st.download_button(
                    "Download merged file",
                    data=output_bytes,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Processing failed: {e}")

card(
    "Notes",
    """
    <div class="small-muted">
      The workflow keeps all other sheets from the uploaded gradebook.
      <br>The selected sheet receives Hours + Email merge and Verdict formula injection.
    </div>
    """,
    muted=True,
)
